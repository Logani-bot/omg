#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
OMG Phase 1 — Coin Top50 데일리 감시 (5년 '고가' + 사이클 고점 규칙 · 엑셀 출력)
===============================================================================

목표 (이번 버전)
- CoinGecko API로 시총 Top 50 목록 수집 (심볼 풀)
- Binance 1d klines의 '고가(high)' 시계열(최근 5년)에 사이클 고점 규칙 적용
  - 규칙: 최저점(L) 이후 +98.5% 상승 시 사이클 시작(mode="high", H = p로 교체)
  - high 모드에서 -44% 이탈 시 mode="wait"로 전환(단, H는 유지)
  - wait 모드에서는 p ≥ 1.985 × L일 때만 사이클 재시작 및 H = p 교체 (직접 신고가만으로는 갱신하지 않음)
  - none 모드에서는 p ≥ 1.985 × L 로 시작하거나, 과거 H가 있고 p > H면 즉시 신고가로 H = p, mode="high"
- 각 코인에 대해 H 기준 매수선(B1~B7)·손절선(Stop) 산출
- 엑셀(xlsx)로 저장 (CONFIG에 예산 항목 포함 → 배분 금액 자동 산출)

의존 라이브러리
- requests
- openpyxl
    pip install requests openpyxl

Author: GPT
Version: 1.4.1
"""
from __future__ import annotations
import time
import pathlib
import datetime as dt
from typing import Any, Dict, List

import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ====== CONFIG ======
VS_CURRENCY = "usd"           # 가격 통화 (usd 권장)
TOP_N = 50                     # 코인 상위 N개 (심볼 풀)
YEARS = 5                      # 5년 윈도우
REQUEST_SLEEP_SEC = 0.8        # API Rate Limit 여유 시간
TIMEOUT_SEC = 20

# 예산 설정 (엑셀 CONFIG 시트에 기본값으로 기록됨)
DEFAULT_BUDGET_PER_ASSET = 1000.0  # 각 자산당 배정 예산(USD)
ALLOC_PCTS = [10, 10, 10, 10, 20, 20, 20]  # 1~7차 매수 비중(%)

# 출력 경로
OUTPUT_DIR = pathlib.Path("./output")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
XLSX_PATH = OUTPUT_DIR / "omg_phase1_levels.xlsx"

# ====== Endpoints ======
CG_BASE = "https://api.coingecko.com/api/v3"
URL_TOP = f"{CG_BASE}/coins/markets"
BINANCE_BASE = "https://api.binance.com"
URL_KLINES = f"{BINANCE_BASE}/api/v3/klines"
URL_EXCHANGE_INFO = f"{BINANCE_BASE}/api/v3/exchangeInfo"

# ====== HTTP (간단 재시도) ======
def http_get(url: str, params: Dict[str, Any]) -> Any:
    for attempt in range(5):
        try:
            resp = requests.get(url, params=params, timeout=TIMEOUT_SEC)
            if resp.status_code == 200:
                return resp.json()
            # 가벼운 백오프
            time.sleep(1 + attempt)
        except requests.RequestException:
            time.sleep(1 + attempt)
    raise RuntimeError(f"GET failed: {url} params={params}")

# ====== Symbol Pool (CoinGecko) ======
def get_top_coins(vs_currency: str = VS_CURRENCY, top_n: int = TOP_N) -> List[Dict[str, Any]]:
    params = {
        "vs_currency": vs_currency,
        "order": "market_cap_desc",
        "per_page": top_n,
        "page": 1,
        "price_change_percentage": "24h",
        "locale": "en",
    }
    data = http_get(URL_TOP, params)
    coins: List[Dict[str, Any]] = []
    for i, row in enumerate(data, start=1):
        coins.append({
            "rank": i,
            "id": row.get("id"),
            "symbol": (row.get("symbol") or "").upper(),
            "name": row.get("name"),
            "market_cap": row.get("market_cap"),
            "current_price": row.get("current_price"),
        })
    return coins

# ====== Exclusions: 파생 / wrapped / 브리지 ======
EXCLUDE_SYMBOLS = {"WBTC", "WETH", "WBETH", "STETH", "WSTETH", "WEETH"}
EXCLUDE_IDS = {"wrapped-bitcoin", "weth", "wrapped-beacon-eth", "staked-ether", "wsteth", "weeth"}
EXCLUDE_NAME_KEYWORDS = {"wrapped", "bridge", "wbtc", "weth", "steth", "wsteth", "weeth"}

def filter_symbol_pool(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    filtered: List[Dict[str, Any]] = []
    for r in rows:
        sym = (r.get("symbol") or "").upper()
        cid = (r.get("id") or "").lower()
        name = (r.get("name") or "").lower()
        if sym in EXCLUDE_SYMBOLS:
            continue
        if cid in EXCLUDE_IDS:
            continue
        if any(k in name for k in EXCLUDE_NAME_KEYWORDS):
            continue
        filtered.append(r)
    return filtered

# ====== Binance (USDT 페어만 사용, A안) ======
def get_binance_usdt_symbol_set() -> set[str]:
    """Binance 스팟에서 상태 TRADING 이고 quoteAsset=USDT 인 심볼 집합."""
    data = http_get(URL_EXCHANGE_INFO, {})
    symbols: set[str] = set()
    for s in data.get("symbols", []):
        try:
            if s.get("status") == "TRADING" and s.get("quoteAsset") == "USDT":
                symbols.add(s.get("symbol"))
        except Exception:
            pass
    return symbols


def map_to_binance_symbol(symbol_upper: str, valid_usdt_symbols: set[str]) -> str | None:
    """CoinGecko 심볼 → Binance USDT 페어. 존재하면 <SYMBOL>USDT, 없으면 None."""
    cand = f"{symbol_upper}USDT"
    return cand if cand in valid_usdt_symbols else None


def get_binance_1d_highs_5y(binance_symbol: str) -> List[float]:
    """최근 YEARS(5)년 구간의 일봉 'high' 시퀀스 반환."""
    now_ms = int(dt.datetime.now(dt.UTC).timestamp() * 1000)
    start_dt = dt.datetime.now(dt.UTC) - dt.timedelta(days=365 * YEARS)
    start_ms = int(start_dt.timestamp() * 1000)

    highs: List[float] = []
    cur_start = start_ms
    while True:
        params = {
            "symbol": binance_symbol,
            "interval": "1d",
            "startTime": cur_start,
            "endTime": now_ms,
            "limit": 1000,
        }
        data = http_get(URL_KLINES, params)
        if not isinstance(data, list) or not data:
            break
        # kline: [openTime, open, high, low, close, volume, closeTime, ...]
        for k in data:
            try:
                highs.append(float(k[2]))
            except Exception:
                pass
        last_close_time = int(data[-1][6])
        if last_close_time >= now_ms:
            break
        cur_start = last_close_time + 1
        if len(highs) > 2200:
            break
        time.sleep(0.2)

    if not highs:
        raise RuntimeError(f"No klines for {binance_symbol}")
    return highs

# ====== Core Logic: 사이클 고점 규칙 ======
def compute_cycle_high(prices: List[float]) -> float | None:
    """
    상태: none / wait / high
    - high: p <= 0.56*H → mode='wait' (H 유지, L=p); p > H → H 갱신
    - wait: L=min(L,p);  p >= 1.985*L 일 때만 H=p로 교체 + mode='high' (p>H 만으로는 갱신 금지)
    - none: L=min(L,p);  p >= 1.985*L → H=p, mode='high'; (보조) 과거 H가 있고 p>H → H=p, mode='high'
    입력 p는 일봉의 'high'
    """
    L: float | None = None
    H: float | None = None
    mode = "none"

    for p in prices:  # 과거 → 현재
        if p is None:
            continue

        if mode == "high":
            if H is not None and p <= H * 0.56:  # -44% 이탈 → wait (H 유지)
                mode = "wait"
                L = p
                continue
            if H is None or p > H:
                H = p
            continue

        # mode in {none, wait}
        if L is None or p < L:
            L = p

        if mode == "wait":
            # wait에서는 오직 +98.5%로만 재시작
            if L is not None and p >= L * 1.985:
                H = p  # 무조건 교체
                mode = "high"
            # p > H 만으로는 갱신하지 않음
            continue

        # mode == "none"
        if L is not None and p >= L * 1.985:
            H = p
            mode = "high"
            continue
        # (보조) 과거 H가 있고 직접 신고가
        if H is not None and p > H:
            H = p
            mode = "high"
            continue

    return H


def compute_levels(H: float) -> Dict[str, float]:
    return {
        "B1": round(H * 0.56, 6),  # -44%
        "B2": round(H * 0.52, 6),  # -48%
        "B3": round(H * 0.46, 6),  # -54%
        "B4": round(H * 0.41, 6),  # -59%
        "B5": round(H * 0.35, 6),  # -65%
        "B6": round(H * 0.28, 6),  # -72%
        "B7": round(H * 0.21, 6),  # -79%
        "Stop": round(H * 0.19, 6), # -81%
    }

# ====== Excel Writer ======
def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)


def write_excel(top_rows: List[Dict[str, Any]], level_rows: List[Dict[str, Any]], skipped_rows: List[Dict[str, Any]]):
    wb = Workbook()

    # CONFIG 시트
    ws_conf = wb.active
    ws_conf.title = "CONFIG"
    ws_conf.append(["Key", "Value", "Note"])
    ws_conf.append(["BudgetPerAsset(USD)", DEFAULT_BUDGET_PER_ASSET, "각 자산별 기본 예산"])
    ws_conf.append(["AllocPercents(1~7)", ",".join(map(str, ALLOC_PCTS)), "매수 비중(%)"])
    autosize(ws_conf)

    # COIN_TOP50
    ws_top = wb.create_sheet("COIN_TOP50")
    ws_top.append(["Rank", "Coin ID", "Symbol", "Name", "MarketCap", "CurrentPrice"])
    for r in top_rows:
        ws_top.append([
            r.get("rank"), r.get("id"), r.get("symbol"), r.get("name"),
            r.get("market_cap"), r.get("current_price")
        ])
    autosize(ws_top)

    # LEVELS
    ws_lvl = wb.create_sheet("LEVELS")
    headers = [
        "Date", "Coin ID", "Symbol", "Name", "H(5y_cycle_high)",
        "B1(-44%)", "B2(-48%)", "B3(-54%)", "B4(-59%)", "B5(-65%)", "B6(-72%)", "B7(-79%)", "Stop(-81%)",
        "Alloc%1", "Alloc%2", "Alloc%3", "Alloc%4", "Alloc%5", "Alloc%6", "Alloc%7",
        "BudgetPerAsset(USD)", "AllocAmt1", "AllocAmt2", "AllocAmt3", "AllocAmt4", "AllocAmt5", "AllocAmt6", "AllocAmt7"
    ]
    ws_lvl.append(headers)

    today = dt.datetime.now().strftime("%Y-%m-%d")
    for r in level_rows:
        budget = r.get("BudgetPerAsset", DEFAULT_BUDGET_PER_ASSET)
        alloc_amts = [round(budget * pct / 100.0, 2) for pct in ALLOC_PCTS]
        ws_lvl.append([
            today, r.get("Coin ID"), r.get("Symbol"), r.get("Name"), r.get("H(5y_cycle_high)"),
            r.get("B1(-44%)"), r.get("B2(-48%)"), r.get("B3(-54%)"), r.get("B4(-59%)"), r.get("B5(-65%)"),
            r.get("B6(-72%)"), r.get("B7(-79%)"), r.get("Stop(-81%)"),
            *ALLOC_PCTS,
            budget, *alloc_amts
        ])
    autosize(ws_lvl)

    # SKIPPED
    ws_skip = wb.create_sheet("SKIPPED")
    ws_skip.append(["Rank", "Coin ID", "Symbol", "Name", "Reason"])
    for r in skipped_rows:
        ws_skip.append([r.get("rank"), r.get("id"), r.get("symbol"), r.get("name"), r.get("reason")])
    autosize(ws_skip)

    wb.save(XLSX_PATH)
    return XLSX_PATH

# ====== Main ======
def main():
    print("[1/4] 시총 Top 50 조회…")
    coins = get_top_coins(vs_currency=VS_CURRENCY, top_n=TOP_N)
    print(f" - {len(coins)}개 수집(원본)")
    coins = filter_symbol_pool(coins)
    print(f" - {len(coins)}개(파생·wrapped·브리지 제외 후) 사용")

    print("[2/4] Binance USDT 심볼 목록 로드…")
    valid_usdt_symbols = get_binance_usdt_symbol_set()
    print(f" - 유효 USDT 심볼 {len(valid_usdt_symbols)}개")

    top_rows = coins

    print("[3/4] 각 코인별 5년 '고가' 시계열에 사이클 고점 규칙 적용…")
    level_rows: List[Dict[str, Any]] = []
    skipped_rows: List[Dict[str, Any]] = []

    for idx, c in enumerate(coins, start=1):
        sym_u = c["symbol"]

        # 스테이블 스킵
        if sym_u in {"USDT", "USDC", "USDS", "DAI", "TUSD", "FDUSD", "USDE"}:
            skipped_rows.append({"rank": c["rank"], "id": c["id"], "symbol": sym_u, "name": c["name"], "reason": "stable"})
            print(f" - [{idx:02d}] {sym_u.lower():<6} 스킵(스테이블)")
            time.sleep(REQUEST_SLEEP_SEC)
            continue

        # Binance USDT 페어 매핑
        binance_sym = map_to_binance_symbol(sym_u, valid_usdt_symbols)
        if binance_sym is None:
            skipped_rows.append({"rank": c["rank"], "id": c["id"], "symbol": sym_u, "name": c["name"], "reason": "no USDT pair on Binance"})
            print(f" ! [{idx:02d}] {sym_u.lower():<6} 스킵 — Binance USDT 페어 없음")
            time.sleep(REQUEST_SLEEP_SEC)
            continue

        try:
            highs = get_binance_1d_highs_5y(binance_sym)  # YEARS=5 윈도우
            H = compute_cycle_high(highs)
            if H is None:
                skipped_rows.append({"rank": c["rank"], "id": c["id"], "symbol": sym_u, "name": c["name"], "reason": "no active cycle H in 5y"})
                print(f" ! [{idx:02d}] {sym_u.lower():<6} 5년 구간에 활성 사이클 H 없음 → 스킵")
            else:
                lvl = compute_levels(H)
                level_rows.append({
                    "Coin ID": c["id"],
                    "Symbol": sym_u,
                    "Name": c["name"],
                    "H(5y_cycle_high)": round(H, 6),
                    "B1(-44%)": lvl["B1"],
                    "B2(-48%)": lvl["B2"],
                    "B3(-54%)": lvl["B3"],
                    "B4(-59%)": lvl["B4"],
                    "B5(-65%)": lvl["B5"],
                    "B6(-72%)": lvl["B6"],
                    "B7(-79%)": lvl["B7"],
                    "Stop(-81%)": lvl["Stop"],
                })
                print(f" - [{idx:02d}] {sym_u.lower():<6} H(5y_cycle_high)={H:.6f}")
        except Exception as e:
            skipped_rows.append({"rank": c["rank"], "id": c["id"], "symbol": sym_u, "name": c["name"], "reason": str(e)[:160]})
            print(f" ! [{idx:02d}] {sym_u.lower():<6} 스킵/실패 — {e}")

        time.sleep(REQUEST_SLEEP_SEC)

    print("[4/4] 엑셀 저장…")
    path = write_excel(top_rows, level_rows, skipped_rows)
    print(f" - 저장완료: {path}")


# ====== Debug Helpers (DOGE step-by-step) ======
from typing import Optional

def get_binance_1d_ohlc_5y(binance_symbol: str) -> List[Dict[str, Any]]:
    """최근 YEARS(5)년 구간의 일봉 OHLC 시퀀스 반환 (리스트의 각 원소는 dict)."""
    now_ms = int(dt.datetime.now(dt.UTC).timestamp() * 1000)
    start_dt = dt.datetime.now(dt.UTC) - dt.timedelta(days=365 * YEARS)
    start_ms = int(start_dt.timestamp() * 1000)

    rows: List[Dict[str, Any]] = []
    cur_start = start_ms
    while True:
        params = {
            "symbol": binance_symbol,
            "interval": "1d",
            "startTime": cur_start,
            "endTime": now_ms,
            "limit": 1000,
        }
        data = http_get(URL_KLINES, params)
        if not isinstance(data, list) or not data:
            break
        for k in data:
            try:
                rows.append({
                    "openTime": int(k[0]),
                    "open": float(k[1]),
                    "high": float(k[2]),
                    "low": float(k[3]),
                    "close": float(k[4]),
                    "closeTime": int(k[6]),
                })
            except Exception:
                pass
        last_close_time = int(data[-1][6])
        if last_close_time >= now_ms:
            break
        cur_start = last_close_time + 1
        if len(rows) > 2200:
            break
        time.sleep(0.2)
    if not rows:
        raise RuntimeError(f"No klines for {binance_symbol}")
    return rows


def debug_cycle_for_symbol(symbol: str = "DOGE", limit_days: Optional[int] = 400) -> None:
    """DOGE 등 단일 심볼에 대해 사이클 로직을 단계별로 추적 출력.
    - limit_days: 최근 N일만 요약 출력(내부 계산은 전체 5년)
    - 결과 CSV: output/<symbol>_debug.csv
    """
    binance_sym = f"{symbol.upper()}USDT"
    print(f"[DEBUG] Fetch {binance_sym} 1d OHLC (5y)…")
    ohlc = get_binance_1d_ohlc_5y(binance_sym)

    # 전체 계산용 high 배열
    highs = [row["high"] for row in ohlc]

    # 추적 변수
    L: float | None = None
    H: float | None = None
    mode = "none"

    # CSV 준비
    out_dir = OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)
    csv_path = out_dir / f"{symbol.lower()}_debug.csv"
    import csv
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["date","open","high","low","close","mode","L","H","event"])  # 헤더

        def ts_to_date(ms: int) -> str:
            return dt.datetime.fromtimestamp(ms/1000, tz=dt.UTC).strftime("%Y-%m-%d")

        for i, row in enumerate(ohlc):
            p = row["high"]
            date = ts_to_date(row["closeTime"])  # 일봉 날짜(UTC)
            event = ""

            if mode == "high":
                if H is not None and p <= H * 0.56:
                    mode = "wait"; L = p; event = "→ wait (−44%)"
                elif H is None or p > H:
                    H = p; event = "H↑ (신고가)"
            else:
                # none/wait 공통: L 갱신
                if L is None or p < L:
                    L = p
                if mode == "wait":
                    if L is not None and p >= L * 1.985:
                        H = p; mode = "high"; event = "restart (+98.5%) H=p"
                else:  # none
                    if L is not None and p >= L * 1.985:
                        H = p; mode = "high"; event = "start (+98.5%) H=p"
                    elif H is not None and p > H:
                        H = p; mode = "high"; event = "start (직접 신고가) H=p"

            w.writerow([
                date, row["open"], row["high"], row["low"], row["close"],
                mode, (None if L is None else round(L,6)), (None if H is None else round(H,6)), event
            ])

    print(f"[DEBUG] CSV saved: {csv_path}")

    # 최근 limit_days만 요약 출력
    if limit_days:
        print(f"[DEBUG] 최근 {limit_days}일 요약 (date, high, mode, L, H, event)")
        with open(csv_path, "r", encoding="utf-8") as f:
            lines = f.readlines()[-limit_days:]
            for ln in lines:
                # 간단 요약: 날짜, 고가, 모드, L, H, 이벤트만 출력
                parts = ln.strip().split(",")
                if parts and parts[0] != "date":
                    date, _open, high, _low, _close, mode_s, Ls, Hs, event = parts
                    print(f" {date} | high={high} | mode={mode_s} | L={Ls} | H={Hs} | {event}")


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == "--debug-doge":
        debug_cycle_for_symbol("DOGE", limit_days=120)
    else:
        main()
